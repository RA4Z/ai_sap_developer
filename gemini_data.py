from openpyxl import load_workbook

arquivo_excel = 'Dev Assistant.xlsx'
workbook = load_workbook(arquivo_excel)
planilha = workbook.worksheets[0]
data = []

for linha in planilha.iter_rows(min_row=2, values_only=True):
    data.append(f"input: {linha[0]}")
    data.append(f"output: {linha[1]}")
