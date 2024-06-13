from openpyxl import Workbook, load_workbook


class ExcelHandler:
    def __init__(self, file_name: str):
        self.file_name = file_name
        self.workbook = None
        self.sheet = None

    def create_workbook(self):
        self.workbook = Workbook()

    def load_workbook(self):
        if 'xlsm' in self.file_name:
            self.workbook = load_workbook(filename=self.file_name, keep_vba=True, keep_links=True)
        else:
            self.workbook = load_workbook(filename=self.file_name)

    def create_sheet(self, sheet_name: str):
        self.sheet = self.workbook.create_sheet(title=sheet_name)

    def select_sheet(self, sheet_name: str):
        self.sheet = self.workbook[sheet_name]

    def write_cell(self, row: int, column: int, value):
        self.sheet.cell(row=row, column=column, value=value)

    def get_cell(self, row: int, column: int):
        return self.sheet.cell(row, column).value

    def save_workbook(self):
        self.workbook.save(self.file_name)

    def close_workbook(self):
        self.workbook.close()

    def count_rows(self, column: int):
        max_row = self.sheet.max_row
        for row in range(max_row, 0, -1):
            cell_value = self.sheet.cell(row, column).value
            if cell_value is not None:
                return row
        return 0

    def count_columns(self, row: int):
        max_col = self.sheet.max_column
        for col in range(max_col, 0, -1):
            cell_value = self.sheet.cell(row=row, column=col).value
            if cell_value is not None:
                return col
        return 0

    def clean_data(self, initial_column: int, final_column: int, initial_row: int, final_row: int):
        # Ensurer the end line is not shorter than the start line
        if final_row < initial_row:
            final_row = initial_row

        # Clean all content between the specified range
        for row in range(initial_row, final_row + 1):
            for col in range(initial_column, final_column + 1):
                self.sheet.cell(row=row, column=col).value = None

    def get_column_index(self, wanted_text: str, start_row=1, end_row=None):
        if end_row is None:
            end_row = start_row

        rng = self.sheet[f"A{start_row}:AZ{end_row}"]
        count = 0
        for row in rng:
            for cell in row:
                if wanted_text.lower() in str(cell.value).lower():
                    count += 1
                    if count == 1:
                        return cell.column
        if count == 0:
            print(f"The column '{wanted_text}' was not found!")
            return None

    def sap_write_my_grid(self, my_grid, total_grid_rows: int, start_row: int, start_column: int):
        grid_column = my_grid.ColumnOrder
        rows = total_grid_rows
        cols = my_grid.ColumnCount
        for i in range(-1, rows):
            for c in range(cols):
                text = my_grid.getCellValue(i, grid_column(c))
                self.write_cell(i + start_row + 1, c + start_column, text)
